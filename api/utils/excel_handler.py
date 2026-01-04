# api/utils/excel_handler.py

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exportador de datos a Excel"""
    
    @staticmethod
    def create_workbook(title="Reporte"):
        """Crear libro de Excel con estilos"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limita a 31 caracteres
        return wb, ws
    
    @staticmethod
    def style_header(ws, header_row=1):
        """Aplicar estilos al header"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    @staticmethod
    def auto_adjust_columns(ws):
        """Ajustar ancho de columnas automáticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def export_products(products, company_name):
        """Exportar productos a Excel"""
        wb, ws = ExcelExporter.create_workbook("Productos")
        
        # Headers
        headers = [
            'Código de Barras', 'Código Paquete', 'Nombre', 'Departamento',
            'Stock Unidades', 'Stock Mínimo', 'Es Paquete', 'Unidades por Paquete',
            'Precio Unitario', 'Precio Paquete', 'Precio Bandeja',
            'Tiene IVA', 'IVA %', 'Envase Retornable', 'Estado'
        ]
        ws.append(headers)
        ExcelExporter.style_header(ws)
        
        # Datos
        for product in products:
            row = [
                product.barcode,
                product.barcode_package or '',
                product.name,
                product.department.name if product.department_id else '',
                float(product.stock_units or 0),
                float(product.min_stock or 0),
                'Sí' if product.is_package else 'No',
                product.units_per_package or '',
                float(product.unit_price or 0),
                float(product.package_price) if product.package_price is not None else '',
                float(product.tray_price) if product.tray_price is not None else '',
                # IVA
                'No' if product.is_tax_exempt else 'Sí',
                float(product.variable_tax_rate or 0) if not product.is_tax_exempt else 0,
                'Sí' if product.has_returnable_container else 'No',
                'Activo' if product.is_active else 'Inactivo',
            ]
            ws.append(row)
        
        ExcelExporter.auto_adjust_columns(ws)
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"productos_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_departments(departments, company_name):
        """Exportar departamentos a Excel"""
        wb, ws = ExcelExporter.create_workbook("Departamentos")
        
        headers = ['Nombre', 'Descripción', 'Cantidad Productos', 'Estado']
        ws.append(headers)
        ExcelExporter.style_header(ws)
        
        for dept in departments:
            row = [
                dept.name,
                dept.description or '',
                dept.products.filter(is_active=True).count(),
                'Activo' if dept.is_active else 'Inactivo'
            ]
            ws.append(row)
        
        ExcelExporter.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"departamentos_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def export_clients(clients, company_name):
        """Exportar clientes a Excel"""
        wb, ws = ExcelExporter.create_workbook("Clientes")
        
        headers = [
            'RUT', 'Nombre', 'Apellido', 'Teléfono', 'Email', 'Dirección',
            'Tiene Crédito', 'Límite Crédito', 'Deuda Actual',
            'Tiene Descuento', 'Descuento %', 'Estado'
        ]
        ws.append(headers)
        ExcelExporter.style_header(ws)
        
        for client in clients:
            row = [
                client.rut,
                client.first_name,
                client.last_name,
                client.phone or '',
                client.email or '',
                client.address or '',
                'Sí' if client.has_credit else 'No',
                float(client.credit_limit) if client.credit_limit else 'Ilimitado',
                float(client.current_debt),
                'Sí' if client.has_discount else 'No',
                float(client.discount_percentage) if client.has_discount else 0,
                'Activo' if client.is_active else 'Inactivo'
            ]
            ws.append(row)
        
        ExcelExporter.auto_adjust_columns(ws)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"clientes_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class ExcelImporter:
    """Importador de datos desde Excel"""
    
    @staticmethod
    def import_products(file, company, department_map):
        """Importar productos desde Excel"""
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            products_data = []
            errors = []
            
            # Saltar header (fila 1)
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Validar que tenga al menos los datos básicos
                    if not row[0] or not row[2]:  # Barcode y Nombre
                        errors.append(f"Fila {idx}: Falta código de barras o nombre")
                        continue
                    
                    product_data = {
                        'barcode': str(row[0]).strip(),
                        'barcode_package': str(row[1]).strip() if row[1] else None,
                        'name': str(row[2]).strip(),
                        'department_name': str(row[3]).strip() if row[3] else None,
                        'stock_units': float(row[4]) if row[4] else 0,
                        'min_stock': float(row[5]) if row[5] else 0,
                        'is_package': str(row[6]).lower() in ['sí', 'si', 'yes', 'true', '1'] if row[6] else False,
                        'units_per_package': int(row[7]) if row[7] else None,
                        'price_unit': float(row[8]) if row[8] else 0,
                        'price_package': float(row[9]) if row[9] else None,
                        'price_tray': float(row[10]) if row[10] else None,
                        'is_tax_exempt': str(row[11]).lower() in ['no', 'false', '0'] if row[11] else False,
                        'tax_percentage': float(row[12]) if row[12] else 19,
                        'has_returnable_container': str(row[13]).lower() in ['sí', 'si', 'yes', 'true', '1'] if row[13] else False,
                        'company': company
                    }
                    
                    products_data.append(product_data)
                    
                except Exception as e:
                    errors.append(f"Fila {idx}: Error al procesar - {str(e)}")
                    logger.error(f"Error importando fila {idx}: {str(e)}")
            
            return {
                'success': len(products_data),
                'products': products_data,
                'errors': errors
            }
            
        except Exception as e:
            logger.error(f"Error al importar Excel: {str(e)}")
            return {
                'success': 0,
                'products': [],
                'errors': [f"Error al leer archivo: {str(e)}"]
            }